"""
Virtual Screening Pipeline  (FIXED v2)
========================================
Automates protein preparation, ligand preparation, active-site detection,
Gnina docking, results extraction, and top-100 Excel export.

Changes vs original
-------------------
* NEW  STEP 0 — Protein preparation  (remove waters / HETATM, add H via
                 OpenBabel when available, fall back to pure-PDB stripping).
* FIX  STEP 2 — Active-site detection:
    ROOT-CAUSE FIX: Tier 1 (co-crystal ligand) now always searches the
    **original raw PDB** (``PROTEIN_PDB``), NOT the prepared protein.
    The protein-preparation step intentionally strips all HETATM records,
    so running Tier 1 on the prepared file made it impossible to detect
    the bound ligand.  Tiers 2 and 3 continue to use the prepared protein
    (cleaner structure, no ligand artefacts).
    ``detect_active_site`` now accepts a ``raw_pdb`` keyword argument for
    this purpose, and ``main()`` passes the original PDB for Tier 1 while
    passing the prepared PDB for Tiers 2/3.

    Additional Tier 3 fixes (carried forward from v1):
    1. Manual 3-D dilation fallback when scipy is absent (replaces the
       broken ``probe_expanded = occupied`` zero-radius probe).
    2. Scoring changed to ``vol^1.5 × burial^0.5 × sphericity``.
    3. Minimum-cluster-size filter (50 voxels) before scoring.
* FIX  STEP 4 — Results saved as CSV + Excel (``top100_results.xlsx``)
                 with the original ``ID`` column from the input spreadsheet.

Dependencies
------------
    pip install rdkit pandas tqdm openpyxl scipy

Gnina must be installed and accessible on PATH:
    https://github.com/gnina/gnina

OpenBabel (optional, improves protein-H placement):
    https://openbabel.org  or  conda install -c conda-forge openbabel
"""

import os
import re
import sys
import subprocess
import logging
import tempfile
import shutil
import math
from pathlib import Path

import pandas as pd
from tqdm import tqdm

from rdkit import Chem
from rdkit.Chem import AllChem, SDWriter, rdMolDescriptors
from rdkit.Chem.rdchem import Mol

# ─────────────────────────────────────────────
# Logging Configuration
# ─────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  [%(levelname)s]  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ─────────────────────────────────────────────
# Global Paths  (edit if needed)
# ─────────────────────────────────────────────
INPUT_EXCEL        = "blind_set.xlsx"
PROTEIN_PDB        = "fixed_version.pdb"          # raw / downloaded structure
PREPARED_PROTEIN   = "fixed_version_prepared.pdb"          # output of Step 0
PREPARED_LIGANDS   = "prepared_ligands.sdf"
DOCKING_OUT_DIR    = Path("docking_outputs")
RESULTS_CSV        = "results_summary.csv"
TOP100_EXCEL       = "top100_results.xlsx"        # NEW: best-100 with ID column

GNINA_EXEC             = "gnina"   # or absolute path
BOX_SIZE               = 25        # Å – slightly larger than original for safety
GNINA_EXHAUSTIVENESS   = 8

# ── Speed settings ────────────────────────────────────────────────────────────
# Safe with cnn_scoring=rescore because the CNN is only called once at the
# very end of each job, so parallel workers don't compete for GPU memory.
GNINA_NUM_WORKERS    = 4   # parallel Gnina processes  (try 4–8)
GNINA_CPU_PER_WORKER = 2   # --cpu passed to each Gnina process
#   Total CPU threads used ≈ GNINA_NUM_WORKERS × GNINA_CPU_PER_WORKER
#   e.g. 4 workers × 2 CPUs = 8 threads.  Keep ≤ your logical core count.


# ══════════════════════════════════════════════════════════════════════════════
# STEP 0 ── Protein Preparation  (NEW)
# ══════════════════════════════════════════════════════════════════════════════

# Residue names that are *never* part of the protein and should be removed
_REMOVE_RESIDUES = {
    # water
    "HOH", "WAT", "DOD", "H2O", "TIP", "TIP3",
    # common cryo-protectants / buffer salts
    "SO4", "PO4", "GOL", "EDO", "ACT", "ACE", "DMS", "MPD",
    "PEG", "EOH", "ETH", "IMD", "TRS", "MES", "BME", "DTT",
    "NHE", "NH4", "FMT", "AZI", "IOD", "BR",  "CL",
    # common ions – remove unless you need them as cofactors
    "NA",  "K",   "CA",  "MG",  "ZN",  "MN",  "CU",  "FE",
    "CO",  "NI",  "CD",  "HG",  "PB",
}


def prepare_protein(
    input_pdb: str,
    output_pdb: str,
    keep_hetatm_residues: set[str] | None = None,
    add_hydrogens: bool = True,
) -> str:
    """
    Prepare a raw PDB structure for molecular docking.

    Steps performed
    ---------------
    1.  Remove all water molecules.
    2.  Remove common crystallographic artefacts (salts, cryo-protectants).
    3.  Keep only ATOM / ANISOU records plus any explicitly listed HETATM
        residues (e.g. essential metal cofactors).
    4.  Retain only the first model (NMR / multi-model structures).
    5.  Add polar hydrogens via OpenBabel (``obabel``) when available;
        otherwise write the cleaned PDB as-is and emit a warning.

    Parameters
    ----------
    input_pdb : str
        Path to the raw receptor PDB file.
    output_pdb : str
        Destination path for the prepared PDB.
    keep_hetatm_residues : set[str] | None
        Residue names of HETATM records to *keep* (e.g. metal cofactors).
        ``None`` → keep nothing (safest default for most docking workflows).
    add_hydrogens : bool
        Whether to attempt hydrogen addition via OpenBabel.

    Returns
    -------
    str
        Path to the prepared PDB (``output_pdb``).
    """
    log.info("═" * 60)
    log.info("STEP 0 — Protein Preparation")
    log.info("═" * 60)

    if not os.path.exists(input_pdb):
        log.error(f"Protein PDB not found: {input_pdb}")
        sys.exit(1)

    keep_hetatm_residues = keep_hetatm_residues or set()

    cleaned_lines: list[str] = []
    in_model = False
    model_done = False
    n_water_removed = 0
    n_hetatm_removed = 0
    n_kept = 0

    with open(input_pdb) as fh:
        for raw in fh:
            rec = raw[:6].strip()

            # ── Handle multi-model PDB (keep Model 1 only) ────────────────
            if rec == "MODEL":
                if model_done:
                    break       # second MODEL encountered – stop reading
                in_model = True
                continue        # skip the MODEL record itself
            if rec == "ENDMDL":
                model_done = True
                in_model = False
                continue

            # ── Always keep ATOM and ANISOU records ───────────────────────
            if rec in ("ATOM", "ANISOU"):
                cleaned_lines.append(raw)
                n_kept += 1
                continue

            # ── Filter HETATM records ────────────────────────────────────
            if rec == "HETATM":
                res_name = raw[17:20].strip().upper()
                if res_name in ("HOH", "WAT", "DOD", "H2O", "TIP", "TIP3"):
                    n_water_removed += 1
                    continue
                if res_name in _REMOVE_RESIDUES:
                    n_hetatm_removed += 1
                    continue
                if res_name in keep_hetatm_residues:
                    cleaned_lines.append(raw)
                    n_kept += 1
                else:
                    n_hetatm_removed += 1
                continue

            # ── Keep structural / bookkeeping records ─────────────────────
            if rec in ("SEQRES", "SSBOND", "LINK", "CRYST1", "REMARK",
                       "TER", "CONECT", "END"):
                cleaned_lines.append(raw)

    # Write the cleaned PDB to a temporary file first
    tmp_clean = output_pdb + ".tmp_clean.pdb"
    with open(tmp_clean, "w") as fh:
        fh.writelines(cleaned_lines)
        if not cleaned_lines[-1].startswith("END"):
            fh.write("END\n")

    log.info(
        f"  Cleaned: kept {n_kept} ATOM records; "
        f"removed {n_water_removed} water atoms, "
        f"{n_hetatm_removed} other HETATM atoms."
    )

    # ── Add Hydrogens ────────────────────────────────────────────────────────
    if add_hydrogens and shutil.which("obabel") is not None:
        log.info("  Adding polar hydrogens with OpenBabel…")
        try:
            proc = subprocess.run(
                [
                    "obabel", tmp_clean,
                    "-O", output_pdb,
                    "-p", "7.4",      # protonate at physiological pH
                    "--partialcharge", "gasteiger",
                ],
                capture_output=True, text=True, timeout=120,
            )
            if proc.returncode == 0:
                log.info("  Hydrogen addition successful (OpenBabel).")
            else:
                log.warning(
                    f"  OpenBabel returned code {proc.returncode}. "
                    "Using cleaned PDB without added hydrogens.\n"
                    f"  stderr: {proc.stderr.strip()[:300]}"
                )
                shutil.copy(tmp_clean, output_pdb)
        except (subprocess.TimeoutExpired, FileNotFoundError):
            log.warning("  OpenBabel timed out or failed. Using cleaned PDB.")
            shutil.copy(tmp_clean, output_pdb)
    else:
        shutil.copy(tmp_clean, output_pdb)
        if add_hydrogens:
            log.warning(
                "  OpenBabel not found on PATH — hydrogens NOT added.\n"
                "  Install with:  conda install -c conda-forge openbabel\n"
                "  Docking will still work; Gnina adds H internally."
            )

    os.remove(tmp_clean)
    log.info(f"  Prepared protein written to '{output_pdb}'.")
    return output_pdb


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 ── Ligand Preparation  (unchanged)
# ══════════════════════════════════════════════════════════════════════════════

def prepare_ligands(excel_path: str, output_sdf: str) -> list[dict]:
    """
    Read SMILES from an Excel file, generate 3-D conformers, minimise with
    MMFF94, and write all successful structures to a multi-molecule SDF file.

    Returns
    -------
    list[dict]
        Each entry: 'index', 'smiles', 'mol_name', and 'ligand_id'
        (the value of the 'ID' column in the Excel sheet, if present).
    """
    log.info("═" * 60)
    log.info("STEP 1 — Ligand Preparation")
    log.info("═" * 60)

    if not os.path.exists(excel_path):
        log.error(f"Input file not found: {excel_path}")
        sys.exit(1)

    df = pd.read_excel(excel_path)

    if "SMILES" not in df.columns:
        log.error("Excel file must contain a 'SMILES' column.")
        sys.exit(1)

    has_id_col = "ID" in df.columns
    if not has_id_col:
        log.warning(
            "  'ID' column not found in Excel file — "
            "results will use row index as identifier."
        )

    log.info(f"Loaded {len(df)} compounds from '{excel_path}'.")

    prepared_mols: list[dict] = []

    with SDWriter(output_sdf) as writer:
        for idx, row in tqdm(
            df.iterrows(),
            total=len(df),
            desc="Preparing ligands",
            unit="mol",
            colour="cyan",
        ):
            smiles   = str(row["SMILES"]).strip()
            mol_name = str(row.get("Name", f"compound_{idx}"))
            lig_id   = str(row["ID"]) if has_id_col else str(idx)

            mol = Chem.MolFromSmiles(smiles)
            if mol is None:
                log.warning(f"  [idx={idx}] Invalid SMILES — skipping: {smiles}")
                continue

            mol.SetProp("_Name", mol_name)

            mol_h = Chem.AddHs(mol)

            params = AllChem.ETKDGv3()
            params.randomSeed = 42
            result = AllChem.EmbedMolecule(mol_h, params)

            if result == -1:
                log.warning(f"  [idx={idx}] Embedding failed — skipping: {mol_name}")
                continue

            ff_result = AllChem.MMFFOptimizeMolecule(mol_h, mmffVariant="MMFF94")
            if ff_result == -1:
                log.warning(
                    f"  [idx={idx}] MMFF94 minimisation failed — skipping: {mol_name}"
                )
                continue

            mol_h.SetProp("SMILES",   smiles)
            mol_h.SetProp("MolIndex", str(idx))
            mol_h.SetProp("LigandID", lig_id)   # carry ID into SDF

            writer.write(mol_h)
            prepared_mols.append(
                {"index": idx, "smiles": smiles, "mol_name": mol_name,
                 "ligand_id": lig_id}
            )

    log.info(
        f"Ligand preparation complete — "
        f"{len(prepared_mols)}/{len(df)} molecules saved to '{output_sdf}'."
    )
    return prepared_mols


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 ── Active Site Detection  (FIXED — three-tier strategy)
# ══════════════════════════════════════════════════════════════════════════════
#
#  Tier 1 — Co-crystallised ligand centroid  (fastest, most reliable)
#  Tier 2 — fpocket  (geometric pocket finder, requires fpocket on PATH)
#  Tier 3 — Built-in grid-based cavity detector  (pure Python + NumPy)
#
# FIXES vs original:
#   • Tier 3 — Dilation fallback: when scipy is absent we now perform a
#     manual 3-D box dilation instead of the broken `probe_expanded = occupied`
#     (the original zero-radius probe caused ALL surface crevices to be flooded
#     as exterior, leaving zero cavity voxels).
#   • Tier 3 — Scoring: changed from vol² × burial to
#     vol^1.5 × burial^0.5 × sphericity (=surface/volume ratio of a sphere).
#     This correctly identifies partially-open kinase-type binding grooves.
#   • Tier 3 — Noise filter: clusters smaller than MIN_CLUSTER_VOXELS are
#     discarded before scoring.
#   • Tier 1 — Added logging of all detected HETATM residues so the user
#     can verify which ligand (if any) was chosen.
# ─────────────────────────────────────────────────────────────────────────────

try:
    import numpy as np
    _NUMPY_AVAILABLE = True
except ImportError:
    _NUMPY_AVAILABLE = False

_SOLVENT_RESIDUES = {
    "HOH", "WAT", "DOD", "H2O",
    "SO4", "PO4", "GOL", "EDO",
    "ACT", "ACE", "DMS", "MPD",
    "PEG", "EOH", "ETH", "IMD",
    "TRS", "MES", "BME", "DTT",
    "NHE", "NH4", "FMT", "AZI",
    "IOD", "BR",  "CL",
    # single-atom ions – never a ligand
    "NA",  "K",   "CA",  "MG",  "ZN",  "MN",  "CU",
    "FE",  "CO",  "NI",  "CD",  "HG",  "PB",
}

_VDW = {
    "C": 1.70, "N": 1.55, "O": 1.52, "S": 1.80, "P": 1.80,
    "F": 1.47, "CL": 1.75, "BR": 1.85, "I": 1.98,
    "ZN": 1.22, "FE": 1.26, "MG": 1.73, "CA": 1.97,
}
_VDW_DEFAULT = 1.70

# Minimum cavity-cluster size (voxels) below which a cluster is discarded.
# At 1 Å resolution, 50 voxels ≈ 50 Å³ — too small to dock a drug-like molecule.
MIN_CLUSTER_VOXELS = 50


def _parse_protein_atoms(pdb_path: str) -> list[tuple[float, float, float, float]]:
    """Return [(x, y, z, vdw_radius), …] for all ATOM records."""
    atoms = []
    with open(pdb_path) as fh:
        for line in fh:
            if not line.startswith("ATOM"):
                continue
            try:
                x = float(line[30:38])
                y = float(line[38:46])
                z = float(line[46:54])
            except ValueError:
                continue
            element = line[76:78].strip().upper() if len(line) > 76 else ""
            if not element:
                element = line[12:16].strip().lstrip("0123456789").upper()[:2]
            r = _VDW.get(element, _VDW_DEFAULT)
            atoms.append((x, y, z, r))
    return atoms


def _centroid(coords: list[tuple]) -> tuple[float, float, float]:
    n = len(coords)
    return (sum(c[0] for c in coords) / n,
            sum(c[1] for c in coords) / n,
            sum(c[2] for c in coords) / n)


# ── Tier 1 ────────────────────────────────────────────────────────────────────

def _tier1_cocrystal(pdb_path: str) -> tuple[float, float, float] | None:
    """Return centroid of the largest co-crystallised ligand, or None."""
    from collections import defaultdict
    residue_coords: dict[str, list[tuple[float, float, float]]] = defaultdict(list)

    with open(pdb_path) as fh:
        for line in fh:
            if not line.startswith("HETATM"):
                continue
            res_name = line[17:20].strip().upper()
            if res_name in _SOLVENT_RESIDUES:
                continue
            try:
                x = float(line[30:38])
                y = float(line[38:46])
                z = float(line[46:54])
            except ValueError:
                continue
            # Key: residue-name + chain + seq-number (unique residue identity)
            chain   = line[21:22].strip()
            seq_num = line[22:26].strip()
            key = f"{res_name}_{chain}_{seq_num}"
            residue_coords[key].append((x, y, z))

    if not residue_coords:
        return None

    # Log every detected HETATM residue so the user can verify
    for key, pts in sorted(residue_coords.items()):
        log.info(f"  [Tier 1] HETATM residue found: {key}  ({len(pts)} heavy atoms)")

    # Use the residue with the most heavy atoms (= most likely the bound ligand)
    best_key  = max(residue_coords, key=lambda k: len(residue_coords[k]))
    best_pts  = residue_coords[best_key]
    cx, cy, cz = _centroid(best_pts)
    log.info(
        f"  [Tier 1] Selected ligand: {best_key}  "
        f"({len(best_pts)} heavy atoms)  →  "
        f"centre ({cx:.3f}, {cy:.3f}, {cz:.3f})"
    )
    return cx, cy, cz


# ── Tier 2 ────────────────────────────────────────────────────────────────────

def _tier2_fpocket(pdb_path: str) -> tuple[float, float, float] | None:
    """Run fpocket and return the centroid of the top-ranked pocket, or None."""
    if shutil.which("fpocket") is None:
        log.debug("  [Tier 2] fpocket not found on PATH — skipping.")
        return None

    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_pdb = os.path.join(tmpdir, "receptor.pdb")
        shutil.copy(pdb_path, tmp_pdb)

        try:
            proc = subprocess.run(
                ["fpocket", "-f", tmp_pdb],
                capture_output=True, text=True, timeout=180, cwd=tmpdir,
            )
        except subprocess.TimeoutExpired:
            log.warning("  [Tier 2] fpocket timed out.")
            return None

        if proc.returncode != 0:
            log.warning(f"  [Tier 2] fpocket exited with code {proc.returncode}.")
            return None

        stem      = Path(tmp_pdb).stem
        info_file = Path(tmpdir) / f"{stem}_out" / f"{stem}_info.txt"

        if not info_file.exists():
            log.warning("  [Tier 2] fpocket info file not found.")
            return None

        cx = cy = cz = None
        in_pocket1 = False
        with open(info_file) as fh:
            for line in fh:
                if re.match(r"Pocket\s+1\s*:", line):
                    in_pocket1 = True
                elif re.match(r"Pocket\s+[2-9]", line):
                    break
                if in_pocket1:
                    m = re.search(r"x_centroid\s*:\s*([-\d.]+)", line)
                    if m: cx = float(m.group(1))
                    m = re.search(r"y_centroid\s*:\s*([-\d.]+)", line)
                    if m: cy = float(m.group(1))
                    m = re.search(r"z_centroid\s*:\s*([-\d.]+)", line)
                    if m: cz = float(m.group(1))

        if None in (cx, cy, cz):
            log.warning("  [Tier 2] Could not parse fpocket centroid.")
            return None

        log.info(
            f"  [Tier 2] fpocket top pocket  →  "
            f"centre ({cx:.3f}, {cy:.3f}, {cz:.3f})"
        )
        return cx, cy, cz


# ── Tier 3  (FIXED) ───────────────────────────────────────────────────────────

def _manual_dilate_3d(grid: "np.ndarray", radius_voxels: int) -> "np.ndarray":
    """
    Simple 3-D spherical dilation using nested loops over a (2r+1)³ kernel.
    Used as a scipy-free fallback for probe-radius expansion.
    Much slower than scipy.ndimage.binary_dilation for large proteins, but
    correct – which the original ``probe_expanded = occupied`` was NOT.
    """
    from collections import deque
    result = grid.copy()
    r = radius_voxels
    # Build integer offsets within a sphere of radius r
    offsets = []
    for dx in range(-r, r + 1):
        for dy in range(-r, r + 1):
            for dz in range(-r, r + 1):
                if dx * dx + dy * dy + dz * dz <= r * r:
                    offsets.append((dx, dy, dz))

    shape = grid.shape
    xs, ys, zs = np.where(grid)
    for x, y, z in zip(xs, ys, zs):
        for dx, dy, dz in offsets:
            nx, ny, nz = x + dx, y + dy, z + dz
            if 0 <= nx < shape[0] and 0 <= ny < shape[1] and 0 <= nz < shape[2]:
                result[nx, ny, nz] = True
    return result


def _tier3_grid(
    pdb_path: str,
    resolution: float = 1.0,
    probe_radius: float = 1.4,
) -> tuple[float, float, float] | None:
    """
    Geometric cavity detection via a 3-D occupancy grid and exterior flood-fill.

    FIXES vs original:
    ─────────────────
    1. ``probe_expanded`` is now computed correctly when scipy is absent:
       instead of ``probe_expanded = occupied`` (zero-radius probe → all
       surface crevices flooded as exterior, leaving no cavity voxels), we
       call ``_manual_dilate_3d``.
    2. Scoring: ``vol^1.5 × burial^0.5 × sphericity`` rewards large AND
       deep pockets while avoiding the quadratic bias that suppressed
       partially-open kinase binding grooves.
    3. Noise filter: clusters < ``MIN_CLUSTER_VOXELS`` voxels are discarded.
    """
    if not _NUMPY_AVAILABLE:
        log.warning("  [Tier 3] NumPy not available — cannot run grid detector.")
        return None

    from collections import deque

    atoms = _parse_protein_atoms(pdb_path)
    if not atoms:
        log.warning("  [Tier 3] No ATOM records found.")
        return None

    coords = np.array([(a[0], a[1], a[2]) for a in atoms])
    radii  = np.array([a[3] for a in atoms])

    log.info(
        f"  [Tier 3] Building occupancy grid for {len(atoms)} atoms "
        f"at {resolution} Å resolution…"
    )

    padding = probe_radius + 3.0
    mins = coords.min(axis=0) - padding
    maxs = coords.max(axis=0) + padding
    shape = tuple(int(math.ceil((maxs[i] - mins[i]) / resolution)) + 1
                  for i in range(3))

    occupied = np.zeros(shape, dtype=bool)

    for (x, y, z), r in zip(coords, radii):
        idx  = ((np.array([x, y, z]) - mins) / resolution).astype(int)
        span = int(math.ceil(r / resolution)) + 1
        xi, yi, zi = idx
        x_lo, x_hi = max(0, xi - span), min(shape[0], xi + span + 1)
        y_lo, y_hi = max(0, yi - span), min(shape[1], yi + span + 1)
        z_lo, z_hi = max(0, zi - span), min(shape[2], zi + span + 1)

        gx = np.arange(x_lo, x_hi)
        gy = np.arange(y_lo, y_hi)
        gz = np.arange(z_lo, z_hi)
        gxx, gyy, gzz = np.meshgrid(gx, gy, gz, indexing="ij")
        rx = mins[0] + gxx * resolution - x
        ry = mins[1] + gyy * resolution - y
        rz = mins[2] + gzz * resolution - z
        mask = (rx**2 + ry**2 + rz**2) <= r * r
        occupied[x_lo:x_hi, y_lo:y_hi, z_lo:z_hi] |= mask

    # ── Probe-expanded occupancy for solvent-accessibility test ──────────────
    probe_voxels = max(1, int(math.ceil(probe_radius / resolution)))
    try:
        from scipy.ndimage import binary_dilation, generate_binary_structure
        struct         = generate_binary_structure(3, 1)
        probe_expanded = occupied.copy()
        for _ in range(probe_voxels):
            probe_expanded = binary_dilation(probe_expanded, structure=struct)
        log.debug("  [Tier 3] Probe dilation via scipy.")
    except ImportError:
        # FIX: use manual dilation instead of the broken fallback
        log.debug(
            "  [Tier 3] scipy not available — using manual 3-D dilation "
            "(slower but correct)."
        )
        probe_expanded = _manual_dilate_3d(occupied, probe_voxels)

    # ── Flood-fill exterior from all border voxels ───────────────────────────
    exterior = np.zeros(shape, dtype=bool)
    queue    = deque()

    def _seed(ix, iy, iz):
        if not probe_expanded[ix, iy, iz] and not exterior[ix, iy, iz]:
            exterior[ix, iy, iz] = True
            queue.append((ix, iy, iz))

    for ix in range(shape[0]):
        for iy in range(shape[1]):
            _seed(ix, iy, 0);            _seed(ix, iy, shape[2] - 1)
    for ix in range(shape[0]):
        for iz in range(shape[2]):
            _seed(ix, 0, iz);            _seed(ix, shape[1] - 1, iz)
    for iy in range(shape[1]):
        for iz in range(shape[2]):
            _seed(0, iy, iz);            _seed(shape[0] - 1, iy, iz)

    directions = [(1,0,0),(-1,0,0),(0,1,0),(0,-1,0),(0,0,1),(0,0,-1)]
    while queue:
        ix, iy, iz = queue.popleft()
        for dx, dy, dz in directions:
            nx, ny, nz = ix + dx, iy + dy, iz + dz
            if (0 <= nx < shape[0] and
                0 <= ny < shape[1] and
                0 <= nz < shape[2]):
                _seed(nx, ny, nz)

    # ── Cavity voxels ────────────────────────────────────────────────────────
    cavity_mask = (~occupied) & (~exterior)
    n_cavity    = int(cavity_mask.sum())

    log.info(
        f"  [Tier 3] Grid shape: {shape}  |  "
        f"occupied: {int(occupied.sum())}  |  "
        f"exterior: {int(exterior.sum())}  |  "
        f"cavity voxels: {n_cavity}"
    )

    if n_cavity == 0:
        log.warning(
            "  [Tier 3] No cavity voxels found. "
            "Try increasing resolution or installing scipy."
        )
        return None

    cavity_pts = list(zip(*np.where(cavity_mask)))

    # ── Cluster contiguous cavity voxels (6-connectivity BFS) ─────────────
    visited  = np.zeros(shape, dtype=bool)
    clusters = []

    for seed in cavity_pts:
        seed = tuple(seed)
        if visited[seed]:
            continue
        cluster = []
        bfs = deque([seed])
        visited[seed] = True
        while bfs:
            vx = bfs.popleft()
            cluster.append(vx)
            for dx, dy, dz in directions:
                nb = (vx[0]+dx, vx[1]+dy, vx[2]+dz)
                if (0 <= nb[0] < shape[0] and
                    0 <= nb[1] < shape[1] and
                    0 <= nb[2] < shape[2] and
                    cavity_mask[nb] and not visited[nb]):
                    visited[nb] = True
                    bfs.append(nb)
        clusters.append(cluster)

    # ── FIX: discard noise clusters below minimum size ────────────────────
    clusters = [c for c in clusters if len(c) >= MIN_CLUSTER_VOXELS]
    if not clusters:
        log.warning(
            f"  [Tier 3] No cavity cluster ≥ {MIN_CLUSTER_VOXELS} voxels found."
        )
        return None

    # ── FIX: improved scoring ──────────────────────────────────────────────
    # score = vol^1.5 × burial^0.5 × sphericity
    # sphericity = (π^(1/3) × (6V)^(2/3)) / A  — penalises elongated crevices
    def _score(cluster):
        volume = len(cluster)
        burial_sum = 0.0
        for vx in cluster:
            buried = sum(
                1 for dx, dy, dz in directions
                if (0 <= vx[0]+dx < shape[0] and
                    0 <= vx[1]+dy < shape[1] and
                    0 <= vx[2]+dz < shape[2] and
                    not exterior[(vx[0]+dx, vx[1]+dy, vx[2]+dz)])
            )
            burial_sum += buried / 6.0
        mean_burial = burial_sum / volume

        # Approximate surface area as number of voxels with ≥1 non-cavity neighbor
        vx_set = set(map(tuple, cluster))
        surface = sum(
            1 for vx in cluster
            if any(
                (vx[0]+dx, vx[1]+dy, vx[2]+dz) not in vx_set
                for dx, dy, dz in directions
            )
        )
        # Ideal sphere with same volume would have surface = (36π V²)^(1/3)
        ideal_surface = (36.0 * math.pi * volume * volume) ** (1.0 / 3.0)
        sphericity = ideal_surface / max(surface, 1e-6)

        return (volume ** 1.5) * (mean_burial ** 0.5) * sphericity

    best_cluster = max(clusters, key=_score)
    best_score   = _score(best_cluster)

    pts         = np.array(best_cluster, dtype=float)
    centre_grid = pts.mean(axis=0)
    cx = float(mins[0] + centre_grid[0] * resolution)
    cy = float(mins[1] + centre_grid[1] * resolution)
    cz = float(mins[2] + centre_grid[2] * resolution)

    log.info(
        f"  [Tier 3] Best pocket: {len(best_cluster)} voxels "
        f"(score={best_score:.1f})  |  "
        f"{len(clusters)} valid clusters  →  "
        f"centre ({cx:.3f}, {cy:.3f}, {cz:.3f})"
    )
    return cx, cy, cz


# ── Public entry point ────────────────────────────────────────────────────────

def detect_active_site(
    pdb_path: str,
    raw_pdb: str | None = None,
) -> tuple[float, float, float]:
    """
    Detect the most probable active-site centre (three-tier strategy).

    Parameters
    ----------
    pdb_path : str
        Path to the **prepared** receptor PDB (waters / HETATM removed).
        Used for Tier 2 (fpocket) and Tier 3 (grid cavity), which benefit
        from a clean protein-only structure.
    raw_pdb : str | None
        Path to the **original, unmodified** PDB file that still contains
        the co-crystallised ligand HETATM records.  When provided, Tier 1
        searches this file instead of ``pdb_path``.

        ROOT-CAUSE of the original bug: the protein-preparation step strips
        all HETATM records, so calling Tier 1 on the prepared file made it
        impossible to find the bound ligand.  Passing the raw PDB here
        restores correct Tier 1 behaviour without compromising the cleaned
        structure used for docking.

        If ``None``, ``pdb_path`` is used for all three tiers (legacy
        behaviour — only safe when no protein preparation has been run).

    Tier 1 — Co-crystallised ligand centroid  → searches ``raw_pdb``
    Tier 2 — fpocket                          → searches ``pdb_path``
    Tier 3 — Grid-based cavity detector       → searches ``pdb_path``
    """
    log.info("═" * 60)
    log.info("STEP 2 — Active Site Detection")
    log.info("═" * 60)

    # The prepared PDB must always exist (it is the docking receptor)
    if not os.path.exists(pdb_path):
        log.error(f"Prepared protein PDB not found: {pdb_path}")
        sys.exit(1)

    # For Tier 1 we prefer the raw (un-stripped) PDB so the co-crystal
    # ligand is still present.  Fall back to pdb_path for legacy callers.
    tier1_pdb = raw_pdb if (raw_pdb is not None) else pdb_path

    if raw_pdb is not None:
        if not os.path.exists(raw_pdb):
            log.warning(
                f"  raw_pdb '{raw_pdb}' not found — "
                "Tier 1 will search the prepared PDB instead "
                "(co-crystal ligand will likely NOT be detected)."
            )
            tier1_pdb = pdb_path
        else:
            log.info(
                f"  Tier 1 will search the original PDB: '{tier1_pdb}'\n"
                f"  Tiers 2/3 will use the prepared PDB:  '{pdb_path}'"
            )
    else:
        log.info(
            "  raw_pdb not supplied — all tiers will use the same PDB.\n"
            "  NOTE: if protein preparation has already been run, Tier 1\n"
            "  may fail to detect the co-crystal ligand (HETATM stripped)."
        )

    # ── Tier 1: co-crystallised ligand (raw PDB) ──────────────────────────
    result = _tier1_cocrystal(tier1_pdb)
    if result is not None:
        log.info("  Detection method: Tier 1 (co-crystallised ligand)")
        return result

    log.info("  No co-crystallised ligand found — trying Tier 2 (fpocket)…")

    # ── Tier 2: fpocket (prepared PDB) ───────────────────────────────────
    result = _tier2_fpocket(pdb_path)
    if result is not None:
        log.info("  Detection method: Tier 2 (fpocket)")
        return result

    log.info("  fpocket unavailable or failed — trying Tier 3 (grid cavity)…")

    # ── Tier 3: grid cavity detector (prepared PDB) ───────────────────────
    result = _tier3_grid(pdb_path)
    if result is not None:
        log.info("  Detection method: Tier 3 (geometric grid cavity)")
        return result

    log.error(
        "All three active-site detection tiers failed.\n"
        "  • Provide a PDB with a co-crystallised ligand (set PROTEIN_PDB\n"
        "    to the original downloaded file before preparation), OR\n"
        "  • Install fpocket (https://fpocket.sourceforge.net/), OR\n"
        "  • Ensure NumPy (+ scipy recommended) is installed."
    )
    sys.exit(1)


# ══════════════════════════════════════════════════════════════════════════════
# STEP 3 ── Gnina Docking  (unchanged except passing receptor through prep path)
# ══════════════════════════════════════════════════════════════════════════════

def run_gnina_docking(
    receptor_pdb: str,
    ligands_sdf: str,
    center: tuple[float, float, float],
    box_size: int = BOX_SIZE,
    out_dir: Path = DOCKING_OUT_DIR,
    gnina_exec: str = GNINA_EXEC,
    exhaustiveness: int = GNINA_EXHAUSTIVENESS,
    num_workers: int = GNINA_NUM_WORKERS,
    cpu_per_worker: int = GNINA_CPU_PER_WORKER,
) -> list[dict]:
    """
    Iterate over molecules in a multi-molecule SDF, run Gnina in parallel,
    and collect raw output log strings.
    """
    import concurrent.futures
    import threading

    log.info("═" * 60)
    log.info("STEP 3 — Gnina Docking")
    log.info("═" * 60)

    if shutil.which(gnina_exec) is None:
        log.error(
            f"Gnina executable not found: '{gnina_exec}'\n"
            "\n"
            "  Option A — download the pre-built Linux binary from GitHub:\n"
            "    wget https://github.com/gnina/gnina/releases/latest/download/gnina\n"
            "    chmod +x gnina && sudo mv gnina /usr/local/bin/\n"
            "\n"
            "  Option B — set GNINA_EXEC at the top of this file to the\n"
            "    full path of your gnina binary.\n"
        )
        sys.exit(1)

    try:
        test = subprocess.run(
            [gnina_exec, "--version"],
            capture_output=True, text=True, timeout=15,
        )
        if test.returncode in (0, 1):
            ver_lines = (test.stdout + test.stderr).strip().splitlines()
            ver_str   = ver_lines[0].strip() if ver_lines else "unknown version"
            log.info(f"  Gnina found: {ver_str}")
        else:
            log.warning(f"  Gnina smoke-test returned code {test.returncode}.")
    except subprocess.TimeoutExpired:
        log.warning("  Gnina smoke-test timed out — continuing anyway.")

    out_dir.mkdir(parents=True, exist_ok=True)
    cx, cy, cz = center

    supplier  = Chem.SDMolSupplier(ligands_sdf, removeHs=False)
    molecules = [mol for mol in supplier if mol is not None]

    if not molecules:
        log.error(f"No valid molecules found in '{ligands_sdf}'.")
        sys.exit(1)

    log.info(
        f"{len(molecules)} ligands queued for docking  "
        f"({num_workers} workers × {cpu_per_worker} CPUs each)."
    )

    _debug_saved = threading.Event()

    def _dock_one(mol) -> dict:
        mol_name  = mol.GetProp("_Name")    if mol.HasProp("_Name")    else "unknown"
        smiles    = mol.GetProp("SMILES")   if mol.HasProp("SMILES")   else ""
        mol_index = mol.GetProp("MolIndex") if mol.HasProp("MolIndex") else "-1"
        lig_id    = mol.GetProp("LigandID") if mol.HasProp("LigandID") else mol_index

        # Include mol_index in filename to avoid collisions between workers
        tmp_sdf      = out_dir / f"{mol_name}_{mol_index}_input.sdf"
        out_sdf_path = out_dir / f"{mol_name}_docked.sdf"

        with SDWriter(str(tmp_sdf)) as w:
            w.write(mol)

        cmd = [
            gnina_exec,
            "--receptor",       receptor_pdb,
            "--ligand",         str(tmp_sdf),
            "--out",            str(out_sdf_path),
            "--center_x",       f"{cx:.4f}",
            "--center_y",       f"{cy:.4f}",
            "--center_z",       f"{cz:.4f}",
            "--size_x",         str(box_size),
            "--size_y",         str(box_size),
            "--size_z",         str(box_size),
            "--exhaustiveness", str(exhaustiveness),
            "--num_modes",      "9",
            "--cnn_scoring",    "rescore",
            "--cpu",            str(cpu_per_worker),
        ]

        raw_log = "TIMEOUT"
        try:
            proc = subprocess.run(
                cmd, capture_output=True, text=True, timeout=300,
            )
            raw_log = proc.stdout + proc.stderr
            if proc.returncode != 0:
                log.warning(
                    f"  Gnina non-zero exit for '{mol_name}': {proc.returncode}"
                )
        except FileNotFoundError:
            log.error(f"Gnina executable not found at '{gnina_exec}'.")
            sys.exit(1)
        except subprocess.TimeoutExpired:
            log.warning(f"  Docking timed out for '{mol_name}' — skipping.")

        # Save debug log for the very first completed job only
        if not _debug_saved.is_set():
            _debug_saved.set()
            debug_path = out_dir / "gnina_debug.txt"
            try:
                with open(debug_path, "w") as dbg:
                    dbg.write(f"=== Command ===\n{' '.join(cmd)}\n\n")
                    dbg.write(f"=== Return code: {proc.returncode} ===\n\n")
                    dbg.write(f"=== STDOUT ===\n{proc.stdout}\n\n")
                    dbg.write(f"=== STDERR ===\n{proc.stderr}\n")
                log.info(f"  Raw gnina output saved to '{debug_path}'.")
            except Exception:
                pass

        tmp_sdf.unlink(missing_ok=True)

        return {
            "mol_name":    mol_name,
            "smiles":      smiles,
            "index":       mol_index,
            "ligand_id":   lig_id,
            "output_path": str(out_sdf_path),
            "raw_log":     raw_log,
        }

    docking_results: list[dict] = []

    with concurrent.futures.ThreadPoolExecutor(max_workers=num_workers) as executor:
        futures = {executor.submit(_dock_one, mol): mol for mol in molecules}
        for future in tqdm(
            concurrent.futures.as_completed(futures),
            total=len(molecules),
            desc="Docking ligands",
            unit="mol",
            colour="green",
        ):
            try:
                docking_results.append(future.result())
            except Exception as exc:
                mol = futures[future]
                name = mol.GetProp("_Name") if mol.HasProp("_Name") else "unknown"
                log.warning(f"  Worker exception for '{name}': {exc}")

    log.info(f"Docking complete — {len(docking_results)} jobs finished.")
    return docking_results


# ══════════════════════════════════════════════════════════════════════════════
# STEP 4 ── Results Extraction + Top-100 Excel Export  (FIXED)
# ══════════════════════════════════════════════════════════════════════════════

# Header keyword patterns
_HDR_AFFINITY     = re.compile(r"affinity",       re.IGNORECASE)
_HDR_CNN_SCORE    = re.compile(r"cnn.?score",     re.IGNORECASE)
_HDR_CNN_AFFINITY = re.compile(r"cnn.?affinity",  re.IGNORECASE)

_KV_AFFINITY      = re.compile(r"affinity\s*[=:]\s*([-\d.]+)",   re.IGNORECASE)
_KV_CNN_SCORE     = re.compile(r"cnn.?score\s*[=:]\s*([\d.]+)",  re.IGNORECASE)
_KV_CNN_AFFINITY  = re.compile(r"cnn.?affinity\s*[=:]\s*([\d.]+)", re.IGNORECASE)


def _split_pipe_row(line: str) -> list[str]:
    parts = [p.strip() for p in line.split("|")]
    return [p for p in parts if p]


def _parse_gnina_log(raw_log: str) -> dict[str, float | None]:
    """
    Parse Gnina stdout to extract mode-1 scores.

    Gnina v1.3.2 produces this format:
        mode |  affinity  |  intramol  |    CNN     |   CNN
             | (kcal/mol) | (kcal/mol) | pose score | affinity
        -----+------------+------------+------------+----------
            1       -8.32       -0.34       0.7574      6.438

    Key quirks handled here:
    - The column headers span TWO lines (affinity on line 1, CNN scores on line 2).
    - Data rows are space-separated, NOT pipe-separated.
    - The separator line starts with "---".
    """
    scores: dict[str, float | None] = {
        "vina_affinity": None,
        "cnn_score":     None,
        "cnn_affinity":  None,
    }

    if not raw_log or raw_log == "TIMEOUT":
        return scores

    lines = raw_log.splitlines()

    # ── Strategy 1: find the "-----+" separator, then parse the next line ─────
    # This is the most reliable approach for gnina's space-delimited table.
    for i, line in enumerate(lines):
        if line.strip().startswith("---") and "+" in line:
            # The data starts on the next non-blank line
            for data_line in lines[i + 1:]:
                stripped = data_line.strip()
                if not stripped:
                    continue
                parts = stripped.split()
                # Expect: mode affinity intramol cnn_pose_score cnn_affinity
                # i.e. at least 5 whitespace-separated tokens, first = "1"
                if len(parts) >= 5 and parts[0] == "1":
                    try:
                        scores["vina_affinity"] = float(parts[1])
                        scores["cnn_score"]     = float(parts[3])
                        scores["cnn_affinity"]  = float(parts[4])
                        return scores
                    except (IndexError, ValueError):
                        pass
                break   # only want mode 1; stop after first data row
            break

    # ── Strategy 2: key=value fallback (older gnina versions) ─────────────────
    m_aff = _KV_AFFINITY.search(raw_log)
    m_cs  = _KV_CNN_SCORE.search(raw_log)
    m_ca  = _KV_CNN_AFFINITY.search(raw_log)
    if m_aff or m_cs or m_ca:
        if m_aff: scores["vina_affinity"] = float(m_aff.group(1))
        if m_cs:  scores["cnn_score"]     = float(m_cs.group(1))
        if m_ca:  scores["cnn_affinity"]  = float(m_ca.group(1))
        return scores

    # ── Strategy 3: pipe-delimited table (some gnina builds) ──────────────────
    for i, line in enumerate(lines):
        if _HDR_AFFINITY.search(line) and _HDR_CNN_SCORE.search(line):
            for data_line in lines[i + 1:]:
                if not data_line.strip() or data_line.strip().startswith("-"):
                    continue
                parts = _split_pipe_row(data_line)
                if not parts or parts[0] != "1":
                    continue
                try:
                    scores["vina_affinity"] = float(parts[1])
                    scores["cnn_score"]     = float(parts[-2])
                    scores["cnn_affinity"]  = float(parts[-1])
                    return scores
                except (IndexError, ValueError):
                    break
            break

    return scores


def extract_results(
    docking_results: list[dict],
    input_excel: str = INPUT_EXCEL,
    output_csv: str = RESULTS_CSV,
    output_excel: str = TOP100_EXCEL,
    top_n: int = 100,
) -> pd.DataFrame:
    """
    Parse Gnina logs, build a results DataFrame, save as CSV, and write the
    top-``top_n`` compounds (by CNN Affinity, then Vina Affinity) to an Excel
    sheet that includes the original ``ID`` column from ``input_excel``.

    Parameters
    ----------
    docking_results : list[dict]
        Output of :func:`run_gnina_docking`.
    input_excel : str
        Path to the original ligand Excel file  (must contain an ``ID`` column).
    output_csv : str
        Destination CSV for the full results table.
    output_excel : str
        Destination Excel file for the top-``top_n`` results WITH the ID column.
    top_n : int
        Number of top-ranked compounds to export  (default 100).

    Returns
    -------
    pd.DataFrame
        Full results table (all compounds, sorted).
    """
    log.info("═" * 60)
    log.info("STEP 4 — Results Extraction & Export")
    log.info("═" * 60)

    # ── Load original Excel to recover the ID column ─────────────────────────
    id_map: dict[str, str] = {}   # mol_index → ID
    if os.path.exists(input_excel):
        orig_df = pd.read_excel(input_excel)
        if "ID" in orig_df.columns:
            for row_idx, row in orig_df.iterrows():
                id_map[str(row_idx)] = str(row["ID"])
        else:
            log.warning(
                f"  'ID' column not found in '{input_excel}'. "
                "The ID column in the output will be populated from LigandID."
            )
    else:
        log.warning(
            f"  Original Excel '{input_excel}' not found — "
            "ID column will be populated from embedded LigandID values."
        )

    # ── Build results DataFrame ───────────────────────────────────────────────
    rows = []
    for entry in docking_results:
        parsed = _parse_gnina_log(entry["raw_log"])

        # ── Fallback: read scores from output SDF properties ─────────────────
        # Gnina embeds scores as SD properties (minimizedAffinity, CNNscore,
        # CNNaffinity) in the output SDF.  This is a reliable fallback when
        # the stdout log is empty (e.g. captured before writing completes).
        if None in parsed.values():
            sdf_path = entry.get("output_path", "")
            if sdf_path and os.path.exists(sdf_path):
                try:
                    suppl = Chem.SDMolSupplier(sdf_path, removeHs=False)
                    first = next((m for m in suppl if m is not None), None)
                    if first is not None:
                        def _prop(mol, *names):
                            for n in names:
                                if mol.HasProp(n):
                                    try: return float(mol.GetProp(n))
                                    except ValueError: pass
                            return None
                        if parsed["vina_affinity"] is None:
                            parsed["vina_affinity"] = _prop(
                                first, "minimizedAffinity", "affinity",
                                "docking_score", "Affinity")
                        if parsed["cnn_score"] is None:
                            parsed["cnn_score"] = _prop(
                                first, "CNNscore", "cnn_score", "CNN_score")
                        if parsed["cnn_affinity"] is None:
                            parsed["cnn_affinity"] = _prop(
                                first, "CNNaffinity", "cnn_affinity",
                                "CNN_affinity")
                except Exception:
                    pass

        # Resolve ID: prefer original Excel lookup, fall back to embedded value
        mol_index = str(entry["index"])
        lig_id    = id_map.get(mol_index, entry.get("ligand_id", mol_index))

        row = {
            "ID":              lig_id,          # ← required output column
            "compound_index":  entry["index"],
            "mol_name":        entry["mol_name"],
            "smiles":          entry["smiles"],
            "vina_affinity":   parsed["vina_affinity"],
            "cnn_score":       parsed["cnn_score"],
            "cnn_affinity":    parsed["cnn_affinity"],
            "output_sdf":      entry["output_path"],
        }
        rows.append(row)

    df = pd.DataFrame(rows)

    # ── Sort: CNN Affinity descending, then Vina Affinity ascending ───────────
    df.sort_values(
        by=["cnn_affinity", "vina_affinity"],
        ascending=[False, True],
        inplace=True,
        na_position="last",
    )

    # ── Save full results as CSV ──────────────────────────────────────────────
    df.to_csv(output_csv, index=False)
    log.info(f"Full results saved to '{output_csv}' ({len(df)} entries).")

    # ── Export top-N to Excel ─────────────────────────────────────────────────
    successful = df.dropna(subset=["cnn_affinity"])
    top_df     = successful.head(top_n).copy()

    # Ensure 'ID' is the first column
    cols = ["ID"] + [c for c in top_df.columns if c != "ID"]
    top_df = top_df[cols]

    top_df.to_excel(output_excel, index=False, sheet_name="Top100")

    # Auto-fit column widths for readability
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        wb = openpyxl.load_workbook(output_excel)
        ws = wb.active
        for col_cells in ws.columns:
            max_len = max(
                (len(str(cell.value)) if cell.value is not None else 0)
                for cell in col_cells
            )
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = (
                min(max_len + 4, 60)
            )
        wb.save(output_excel)
    except Exception as exc:
        log.debug(f"  Column auto-fit skipped: {exc}")

    log.info(
        f"Top {len(top_df)} compounds saved to '{output_excel}' "
        f"(sheet: 'Top100')."
    )

    # ── Console summary ───────────────────────────────────────────────────────
    log.info(f"Successfully scored: {len(successful)}/{len(df)} compounds.")
    if not successful.empty:
        best = successful.iloc[0]
        log.info(
            f"  Top compound  ID={best['ID']}  |  "
            f"CNN Affinity: {best['cnn_affinity']:.3f}  |  "
            f"CNN Score: {best['cnn_score']:.3f}  |  "
            f"Vina: {best['vina_affinity']:.3f} kcal/mol"
        )

    return df


# ══════════════════════════════════════════════════════════════════════════════
# Recovery helper — re-score from existing docked SDF files
# ══════════════════════════════════════════════════════════════════════════════

def recover_results(
    ligands_sdf: str = PREPARED_LIGANDS,
    out_dir: Path = DOCKING_OUT_DIR,
    input_excel: str = INPUT_EXCEL,
    output_csv: str = RESULTS_CSV,
    output_excel: str = TOP100_EXCEL,
    top_n: int = 100,
) -> pd.DataFrame:
    """
    Re-extract results from already-docked SDF files without re-running Gnina.

    Use this when docking completed successfully but Step 4 scored 0 compounds
    (e.g. due to a log-parsing bug).  Reads compound list from the prepared
    ligands SDF, locates each ``<mol_name>_docked.sdf`` in ``out_dir``, and
    reads scores from the SDF properties that Gnina embeds in every output file.

    Usage
    -----
        python virtual_screening_fast.py --recover
    """
    log.info("═" * 60)
    log.info("RECOVERY — Re-scoring from existing docked SDF files")
    log.info("═" * 60)

    supplier  = Chem.SDMolSupplier(ligands_sdf, removeHs=False)
    molecules = [mol for mol in supplier if mol is not None]
    log.info(f"  {len(molecules)} ligands found in '{ligands_sdf}'.")

    docking_results = []
    missing = 0
    for mol in tqdm(molecules, desc="Reading docked SDFs", unit="mol", colour="cyan"):
        mol_name  = mol.GetProp("_Name")    if mol.HasProp("_Name")    else "unknown"
        smiles    = mol.GetProp("SMILES")   if mol.HasProp("SMILES")   else ""
        mol_index = mol.GetProp("MolIndex") if mol.HasProp("MolIndex") else "-1"
        lig_id    = mol.GetProp("LigandID") if mol.HasProp("LigandID") else mol_index

        out_sdf = out_dir / f"{mol_name}_docked.sdf"
        if not out_sdf.exists():
            missing += 1
            raw_log = "TIMEOUT"
        else:
            # Build a fake raw_log by reading the docked SDF's first pose
            raw_log = ""
            try:
                suppl2 = Chem.SDMolSupplier(str(out_sdf), removeHs=False)
                first  = next((m for m in suppl2 if m is not None), None)
                if first is not None:
                    aff  = first.GetProp("minimizedAffinity") if first.HasProp("minimizedAffinity") else ""
                    cs   = first.GetProp("CNNscore")          if first.HasProp("CNNscore")          else ""
                    ca   = first.GetProp("CNNaffinity")        if first.HasProp("CNNaffinity")        else ""
                    # Reconstruct a minimal gnina-style table so _parse_gnina_log works
                    raw_log = (
                        "mode |  affinity  |  intramol  |    CNN     |   CNN\n"
                        "     | (kcal/mol) | (kcal/mol) | pose score | affinity\n"
                        "-----+------------+------------+------------+----------\n"
                        f"    1    {aff}       0.00    {cs}    {ca}\n"
                    )
            except Exception:
                pass

        docking_results.append({
            "mol_name":    mol_name,
            "smiles":      smiles,
            "index":       mol_index,
            "ligand_id":   lig_id,
            "output_path": str(out_dir / f"{mol_name}_docked.sdf"),
            "raw_log":     raw_log,
        })

    if missing:
        log.warning(f"  {missing} docked SDF files were missing (will appear as unscored).")

    return extract_results(
        docking_results=docking_results,
        input_excel=input_excel,
        output_csv=output_csv,
        output_excel=output_excel,
        top_n=top_n,
    )


# ══════════════════════════════════════════════════════════════════════════════
# Main Pipeline
# ══════════════════════════════════════════════════════════════════════════════

def main() -> pd.DataFrame:
    log.info("╔" + "═" * 58 + "╗")
    log.info("║        VIRTUAL SCREENING PIPELINE  —  START           ║")
    log.info("╚" + "═" * 58 + "╝")

    # ── Step 0: Prepare Protein  (NEW) ──────────────────────────────────────
    prepared_receptor = prepare_protein(
        input_pdb=PROTEIN_PDB,
        output_pdb=PREPARED_PROTEIN,
        # To keep e.g. a zinc cofactor:  keep_hetatm_residues={"ZN"}
        keep_hetatm_residues=None,
        add_hydrogens=True,
    )

    # ── Step 1: Prepare Ligands ──────────────────────────────────────────────
    prepared = prepare_ligands(
        excel_path=INPUT_EXCEL,
        output_sdf=PREPARED_LIGANDS,
    )

    if not prepared:
        log.error("No ligands were prepared successfully. Exiting.")
        sys.exit(1)

    # ── Step 2: Detect Active Site ───────────────────────────────────────────
    # CRITICAL: pass the ORIGINAL (raw) PDB as raw_pdb so that Tier 1 can
    # find the co-crystallised ligand, which was removed from PREPARED_PROTEIN
    # by the protein-preparation step.  Tiers 2 and 3 automatically fall back
    # to the prepared protein (pdb_path) for a cleaner cavity search.
    center = detect_active_site(
        pdb_path=prepared_receptor,   # prepared protein → Tiers 2 & 3
        raw_pdb=PROTEIN_PDB,          # original PDB     → Tier 1 (co-crystal)
    )

    # ── Step 3: Run Gnina Docking ────────────────────────────────────────────
    docking_results = run_gnina_docking(
        receptor_pdb=prepared_receptor,
        ligands_sdf=PREPARED_LIGANDS,
        center=center,
        box_size=BOX_SIZE,
        out_dir=DOCKING_OUT_DIR,
        gnina_exec=GNINA_EXEC,
        exhaustiveness=GNINA_EXHAUSTIVENESS,
        num_workers=GNINA_NUM_WORKERS,
        cpu_per_worker=GNINA_CPU_PER_WORKER,
    )

    # ── Step 4: Extract & Save Results ──────────────────────────────────────
    results_df = extract_results(
        docking_results=docking_results,
        input_excel=INPUT_EXCEL,
        output_csv=RESULTS_CSV,
        output_excel=TOP100_EXCEL,
        top_n=100,
    )

    log.info("╔" + "═" * 58 + "╗")
    log.info("║        VIRTUAL SCREENING PIPELINE  —  DONE            ║")
    log.info("╚" + "═" * 58 + "╝")

    return results_df


# ─────────────────────────────────────────────
if __name__ == "__main__":
    import sys
    if "--recover" in sys.argv:
        recover_results()
    else:
        main()
